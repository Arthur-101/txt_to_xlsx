import re
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Reading the text file:-
with open('65027 XII.txt') as file:
    reader = file.read()

# Defining the patterns to get 'Date', 'School Code', 'School Name' and Region' :-
date_pattern = r"DATE:- (\d{2}/\d{2}/\d{4})"
school_pattern = r"SCHOOL : - (\d+) (.+)"
region_pattern = r"REGION:\s+(\S+)"

# Finding the matches for 'Date', 'School code', 'School name', and 'Region' :-
date_match = re.search(date_pattern, reader)
school_match = re.search(school_pattern, reader)
region_match = re.search(region_pattern, reader)

# Giving variable name to all the data:-
date = date_match.group(1) if date_match else ""
school_code = school_match.group(1) if school_match else ""
school_name = school_match.group(2) if school_match else ""
region = region_match.group(1) if region_match else ""

# Defining the patterns to get all the outro information :-
total_candidates_pattern = r"TOTAL CANDIDATES\s*:\s*(\d+)"
total_pass_pattern = r"TOTAL PASS\s*:\s*(\d+)"
total_comptt_pattern = r"TOTAL COMPTT\.\s*:\s*(\d+)"
total_essential_repeat_pattern = r"TOTAL ESSENTIAL REPEAT\s*:\s*(\d+)"
total_absent_pattern = r"TOTAL ABSENT\s*:\s*(\d+)"

# Finding the matches of above defined pattern in the file:-
total_candidates_match = re.search(total_candidates_pattern, reader)
total_pass_match = re.search(total_pass_pattern, reader)
total_comptt_match = re.search(total_comptt_pattern, reader)
total_essential_repeat_match = re.search(total_essential_repeat_pattern, reader)
total_absent_match = re.search(total_absent_pattern, reader)

# Giving variable names to all the data :-
total_candidates = total_candidates_match.group(1) if total_candidates_match else ""
total_pass = total_pass_match.group(1) if total_pass_match else ""
total_comptt = total_comptt_match.group(1) if total_comptt_match else ""
total_essential_repeat = total_essential_repeat_match.group(1) if total_essential_repeat_match else ""
total_absent = total_absent_match.group(1) if total_absent_match else ""

# Pattern to match the unwanted text (from the top of the string to the line starting with "SCHOOL")
unwanted_pattern = r'DATE:.*?\n.*?-----.*?\n\nSCHOOL.*?\n'

# Remove unwanted text using re.sub
input_string_cleaned = re.sub(unwanted_pattern, '', reader, flags=re.DOTALL)

# Pattern to extract Roll, Gender, and Name
roll_gender_name_pattern = r"(\d+)\s+(\w)\s+([A-Z ]+)"

# Pattern to extract Subject Codes
subject_codes_pattern = r"(\d{3}\s+){5}(\d{3})?"

# Pattern to extract Result
result_pattern = r"\b(PASS|FAIL|COMP|ABST)\b"

# Pattern to extract Marks and Grades for each subject
marks_grades_pattern = r"(\d{3}|AB)\s+([A-Z]\d?)"

# Initializing an empty list to store each student's data as a dictionary
students_data = []

# Initializing variables to keep track of student information
current_student_info = None
current_student_grades = None

# Spliting the input_string by newline characters
lines = reader.strip().split('\n')

# Defining a Function to process and add student data to students_data list
def add_student_data(roll, gender, name, subject_codes, result, marks_grades):
    marks = []
    grades = []
    for mark_grade_tuple in marks_grades:
        mark, grade = mark_grade_tuple[0], mark_grade_tuple[1]
        if mark.isdigit():
            marks.append(int(mark))
        else:
            marks.append(mark)
        grades.append(grade)
    
    if len(subject_codes) < 6 and len(marks) < 6 and len(grades) < 6:
        subject_codes.append(np.NaN)
        marks.append(np.NaN)
        grades.append(np.NaN)

    row_data = {
        'Roll': roll,
        'Gender': gender,
        'Name': name.strip(),
        'Sub_1': subject_codes[0],
        'Marks_1': marks[0],
        'grade_1': grades[0],
        'Sub_2': subject_codes[1],
        'Marks_2': marks[1],
        'grade_2': grades[1],
        'Sub_3': subject_codes[2],
        'Marks_3': marks[2],
        'grade_3': grades[2],
        'Sub_4': subject_codes[3],
        'Marks_4': marks[3],
        'grade_4': grades[3],
        'Sub_5': subject_codes[4],
        'Marks_5': marks[4],
        'grade_5': grades[4],
        'Sub_6': subject_codes[5],
        'Marks_6': marks[5],
        'grade_6': grades[5],
        'Result': result
    }
    
    students_data.append(row_data)

# Iterating through each line to process student data
for line in lines:
    line = line.strip()
    # Checking if the line contains Roll, Gender, and Name
    if re.match(roll_gender_name_pattern, line):
        current_student_info = line
    # Checking if the line contains Marks and Grades
    elif re.match(marks_grades_pattern, line):
        current_student_grades = line
        # Extracting Roll, Gender, and Name
        roll, gender, name = re.search(roll_gender_name_pattern, current_student_info).groups()

        # Extracting Subject Codes and Separating individual subject codes
        subject_codes_string = re.search(subject_codes_pattern, current_student_info).group()
        subject_codes = re.findall(r"\d{3}", subject_codes_string)

        # Extracting Result
        result = re.search(result_pattern, current_student_info).group()

        # Extracting Marks and Grades for each subject
        marks_grades = re.findall(marks_grades_pattern, current_student_grades)

        # Adding student data to students_data list
        add_student_data(roll, gender, name, subject_codes, result, marks_grades)

# Creating the final DataFrame using the list of student dictionaries
df = pd.DataFrame(students_data)

# Reseting the index of the DataFrame
df.reset_index(drop=True, inplace=True)
df = df.fillna("")

# Printing the DataFrame
print(df)

# Creating an Excel Workbook
workbook = Workbook()
sheet = workbook.active

# Adding additional information (Date, School Code, School Name, Region) to the Excel file
sheet['A1'] = 'Date:-'
sheet['A2'] = 'School Code'
sheet['H1'] = 'Region:'

sheet['B1'] = date
sheet['B2'] = school_code
sheet['I1'] = region

start_column = 'C'
end_column = 'G'
merged_cell = sheet[start_column + '3']
merged_cell.alignment = Alignment(horizontal='center', vertical='center')
sheet.merge_cells(f'{start_column}3:{end_column}3')           # Merging cells for the school_name
sheet[f'{start_column}3'] = school_name

# Writing DataFrame rows to Excel, starting from the specified row
for row in dataframe_to_rows(df, index=False, header=True, ):
    sheet.append(row)

start_row = 4

# Adding Outro Information :-
end = sheet.max_row
sheet['A'+str(end+2)] = 'Total Candidates :- '
sheet['D'+str(end+2)] = 'Total Absent :- '
sheet['A'+str(end+3)] = 'Total Pass :- '
sheet['D'+str(end+3)] = 'Total Comptt. :-'
sheet['A'+str(end+4)] = 'Total Essential Repeat :- '

sheet['B'+str(end+2)] = total_candidates
sheet['E'+str(end+2)] = total_absent
sheet['B'+str(end+3)] = total_pass
sheet['E'+str(end+3)] = total_comptt
sheet['B'+str(end+4)] = total_essential_repeat

# Defining alignment settings for specific columns
alignment_settings = {
    'A': Alignment(horizontal='left', vertical='center'),
    'B': Alignment(horizontal='center', vertical='center'),
    'C': Alignment(horizontal='left', vertical='center'),
    'D': Alignment(horizontal='right', vertical='center'),
    'E': Alignment(horizontal='center', vertical='center'),
    'F': Alignment(horizontal='left', vertical='center'),
    'G': Alignment(horizontal='right', vertical='center'),
    'H': Alignment(horizontal='center', vertical='center'),
    'I': Alignment(horizontal='left', vertical='center'),
    'J': Alignment(horizontal='right', vertical='center'),
    'K': Alignment(horizontal='center', vertical='center'),
    'L': Alignment(horizontal='left', vertical='center'),
    'M': Alignment(horizontal='right', vertical='center'),
    'N': Alignment(horizontal='center', vertical='center'),
    'O': Alignment(horizontal='left', vertical='center'),
    'P': Alignment(horizontal='right', vertical='center'),
    'Q': Alignment(horizontal='center', vertical='center'),
    'R': Alignment(horizontal='left', vertical='center'),
    'S': Alignment(horizontal='right', vertical='center'),
    'T': Alignment(horizontal='center', vertical='center'),
    'U': Alignment(horizontal='left', vertical='center'),
    'V': Alignment(horizontal='center', vertical='center'),
}

# Applying alignment settings to specific columns
for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        column_letter = cell.column_letter
        if column_letter in alignment_settings:
            cell.alignment = alignment_settings[column_letter]

# Saving the Excel file
workbook.save('z_example.xlsx')

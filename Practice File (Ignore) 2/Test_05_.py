import re
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Files :-
open_file_from = '65027 XII.txt'
save_file_at = 'Output_Excel_2.xlsx'
# Reading the text file:-
with open(open_file_from) as file:
    reader = file.read()

# Defining the patterns to get 'Date', 'School Code', 'School Name' and Region' :-
date_pattern = r"DATE:- (\d{2}/\d{2}/\d{4})"
school_pattern = r"SCHOOL : - (\d+) (.+)"
region_pattern = r"REGION:\s+(\S+)"

# Finding the matches for 'Date', 'School code', 'School name', and 'Region' :-
date_match = re.search(date_pattern, reader)
school_match = re.search(school_pattern, reader)
region_match = re.search(region_pattern, reader)

# Giving variable name to all the data:-
date = date_match.group(1) if date_match else ""
school_code = school_match.group(1) if school_match else ""
school_name = school_match.group(2) if school_match else ""
region = region_match.group(1) if region_match else ""

# Defining the patterns to get all the outro information :-
total_candidates_pattern = r"TOTAL CANDIDATES\s*:\s*(\d+)"
total_pass_pattern = r"TOTAL PASS\s*:\s*(\d+)"
total_comptt_pattern = r"TOTAL COMPTT\.\s*:\s*(\d+)"
total_essential_repeat_pattern = r"TOTAL ESSENTIAL REPEAT\s*:\s*(\d+)"
total_absent_pattern = r"TOTAL ABSENT\s*:\s*(\d+)"

# Finding the matches of above defined pattern in the file:-
total_candidates_match = re.search(total_candidates_pattern, reader)
total_pass_match = re.search(total_pass_pattern, reader)
total_comptt_match = re.search(total_comptt_pattern, reader)
total_essential_repeat_match = re.search(total_essential_repeat_pattern, reader)
total_absent_match = re.search(total_absent_pattern, reader)

# Giving variable names to all the data :-
total_candidates = total_candidates_match.group(1) if total_candidates_match else ""
total_pass = total_pass_match.group(1) if total_pass_match else ""
total_comptt = total_comptt_match.group(1) if total_comptt_match else ""
total_essential_repeat = total_essential_repeat_match.group(1) if total_essential_repeat_match else ""
total_absent = total_absent_match.group(1) if total_absent_match else ""

# Pattern to match the unwanted text (from the top of the string to the line starting with "SCHOOL")
unwanted_pattern = r'DATE:.*?\n.*?-----.*?\n\nSCHOOL.*?\n'

# Remove unwanted text using re.sub
input_string_cleaned = re.sub(unwanted_pattern, '', reader, flags=re.DOTALL)

# Pattern to extract Roll, Gender, and Name
roll_gender_name_pattern = r"(\d+)\s+(\w)\s+([A-Z ]+)"

# Pattern to extract Subject Codes
subject_codes_pattern = r"(\d{3}\s+){5}(\d{3})?"

# Pattern to extract Result
result_pattern = r"\b(PASS|FAIL|COMP|ABST)\b"

# Pattern to extract Marks and Grades for each subject
marks_grades_pattern = r"(\d{3}|AB)\s+([A-Z]\d?)"

# Initializing an empty list to store each student's data as a dictionary
students_data = []

# Initializing variables to keep track of student information
current_student_info = None
current_student_grades = None

# Spliting the input_string by newline characters
lines = reader.strip().split('\n')

# Defining a Function to process and add student data to students_data list
def add_student_data(roll, gender, name, subject_codes, result, marks_grades):
    marks = []
    grades = []
    for mark_grade_tuple in marks_grades:
        mark, grade = mark_grade_tuple[0], mark_grade_tuple[1]
        if mark.isdigit():
            marks.append(int(mark))
        else:
            marks.append(mark)
        grades.append(grade)

    if len(subject_codes) < 6 and len(marks) < 6 and len(grades) < 6:
        subject_codes.append(np.NaN)
        marks.append(np.NaN)
        grades.append(np.NaN)

    row_data = {
        'Roll': roll,
        'Gender': gender,
        'Name': name.strip(),
        'Sub_1': subject_codes[0],
        'Marks_1': marks[0],
        'grade_1': grades[0],
        'Sub_2': subject_codes[1],
        'Marks_2': marks[1],
        'grade_2': grades[1],
        'Sub_3': subject_codes[2],
        'Marks_3': marks[2],
        'grade_3': grades[2],
        'Sub_4': subject_codes[3],
        'Marks_4': marks[3],
        'grade_4': grades[3],
        'Sub_5': subject_codes[4],
        'Marks_5': marks[4],
        'grade_5': grades[4],
        'Sub_6': subject_codes[5],
        'Marks_6': marks[5],
        'grade_6': grades[5],
        'Result': result
    }
    
    students_data.append(row_data)

# Iterating through each line to process student data
for line in lines:
    line = line.strip()
    # Checking if the line contains Roll, Gender, and Name
    if re.match(roll_gender_name_pattern, line):
        current_student_info = line
    # Checking if the line contains Marks and Grades
    elif re.match(marks_grades_pattern, line):
        current_student_grades = line
        # Extracting Roll, Gender, and Name
        roll, gender, name = re.search(roll_gender_name_pattern, current_student_info).groups()

        # Extracting Subject Codes and Separating individual subject codes
        subject_codes_string = re.search(subject_codes_pattern, current_student_info).group()
        subject_codes = re.findall(r"\d{3}", subject_codes_string)

        # Extracting Result
        result = re.search(result_pattern, current_student_info).group()

        # Extracting Marks and Grades for each subject
        marks_grades = re.findall(marks_grades_pattern, current_student_grades)

        # Adding student data to students_data list
        add_student_data(roll, gender, name, subject_codes, result, marks_grades)

# Creating the final DataFrame using the list of student dictionaries
df = pd.DataFrame(students_data)

# Reseting the index of the DataFrame
df.reset_index(drop=True, inplace=True)
df = df.fillna("")

# Function to extract numeric value from a string
def extract_numeric(value):
    numeric_part = re.sub(r'\D', '', str(value))
    return int(numeric_part) if numeric_part.isdigit() else 0

# Apply the function to extract numeric values from Marks columns
df['Marks_1'] = df['Marks_1'].apply(extract_numeric)
df['Marks_2'] = df['Marks_2'].apply(extract_numeric)
df['Marks_3'] = df['Marks_3'].apply(extract_numeric)
df['Marks_4'] = df['Marks_4'].apply(extract_numeric)
df['Marks_5'] = df['Marks_5'].apply(extract_numeric)
df['Marks_6'] = df['Marks_6'].apply(extract_numeric)

# Add a new column 'Total_Marks'
df['Total_Marks'] = df['Marks_1'] + df['Marks_2'] + df['Marks_3'] + df['Marks_4'] + df['Marks_5'] + df['Marks_6']

# Calculating and Summing Up the best five marks for each row and adding a new column 'Total Marks (Best 5)'
df['Total Marks (Best 5)'] = df[['Marks_1', 'Marks_2', 'Marks_3', 'Marks_4', 'Marks_5', 'Marks_6']].apply(lambda row: sum(sorted(row, reverse=True)[:5]), axis=1)

# Calculate percentage and add a new column 'Percentage (%)'
total_possible_marks = 500
df['Percentage (%)'] = (df['Total Marks (Best 5)'] / total_possible_marks) * 100
df['Percentage (%)'] = df['Percentage (%)'].apply(lambda x: round(x, 2))

# Extract unique subject codes
subject_codes = df[['Sub_1', 'Sub_2', 'Sub_3', 'Sub_4', 'Sub_5', 'Sub_6']].values.flatten()
subject_codes = np.unique(subject_codes)

# Create a new dataframe
new_columns = ['Roll', 'Gender', 'Name']
for code in subject_codes:
    new_columns.append(code)
    new_columns.append('Grade_' + code)
new_columns.append('Total Marks')                # Add a new column 'Total_Marks'
new_columns.append('Total Marks (Best 5)')         # Add a new column 'Total_Marks (Best 5)'

new_df = pd.DataFrame(columns=new_columns)

# Iterate through rows and populate the new dataframe
for _, row in df.iterrows():
    new_row = [row['Roll'], row['Gender'], row['Name']]
    total_marks = 0
    best_marks = []
    for code in subject_codes:
        matching_index = np.where(row[['Sub_1', 'Sub_2', 'Sub_3', 'Sub_4', 'Sub_5', 'Sub_6']] == code)[0]
        if matching_index.size > 0:
            marks = row['Marks_' + str(matching_index[0] + 1)]
            grade = row['grade_' + str(matching_index[0] + 1)]
            total_marks += marks           # Adding total marks in a list
            best_marks.append(marks)       # Adding total marks (Best 5) in a list
            new_row.append(marks)
            new_row.append(grade)
        else:
            new_row.append(np.NaN)
            new_row.append(np.NaN)

    # Adding all `Total_Marks` and `Total_Marks (Best 5)` in new columns
    best_marks.sort(reverse=True)
    total_marks_best_5 = sum(best_marks[:5])
    new_row.append(total_marks)
    new_row.append(total_marks_best_5)

    new_df.loc[len(new_df)] = new_row

# Calculate percentage and add a new column 'Percentage (%)'
total_possible_marks = 500
new_df['Percentage (%)'] = (new_df['Total Marks (Best 5)'] / total_possible_marks) * 100
new_df['Percentage (%)'] = new_df['Percentage (%)'].apply(lambda x: round(x, 2))

new_df = new_df.fillna('')

# Iterate through columns in the DataFrame to check for empty columns and remove it
for col in new_df.columns:
    if all(new_df[col] == ""):
        new_df.drop(columns=[col], inplace=True)

# Remove columns with empty string as name
columns_to_remove = [col for col in new_df.columns if col == ""]
new_df.drop(columns=columns_to_remove, inplace=True)

# Printing the new_df
new_df

# Creating an Excel Workbook
workbook = Workbook()
sheet = workbook.active
sheet.title = 'Result'

# Adding additional information (Date, School Code, School Name, Region) to the Excel file
sheet['A1'] = 'Date:-'
sheet['A2'] = 'School Code'
sheet['H1'] = 'Region:'

sheet['B1'] = date
sheet['B2'] = school_code
sheet['I1'] = region

start_column = 'C'
end_column = 'G'
merged_cell = sheet[start_column + '3']
merged_cell.alignment = Alignment(horizontal='center', vertical='center')
sheet.merge_cells(f'{start_column}3:{end_column}3')           # Merging cells for the school_name
sheet[f'{start_column}3'] = school_name

# Writing DataFrame rows to Excel, starting from the specified row
for row in dataframe_to_rows(new_df, index=False, header=True, ):
    sheet.append(row)
sheet.append([' '])

start_row = 4

# Adding Outro Information :-
end = sheet.max_row
sheet['A'+str(end+2)] = 'Total Candidates :- '
sheet['D'+str(end+2)] = 'Total Absent :- '
sheet['A'+str(end+3)] = 'Total Pass :- '
sheet['D'+str(end+3)] = 'Total Comptt. :-'
sheet['A'+str(end+4)] = 'Total Essential Repeat :- '

sheet['B'+str(end+2)] = total_candidates
sheet['E'+str(end+2)] = total_absent
sheet['B'+str(end+3)] = total_pass
sheet['E'+str(end+3)] = total_comptt
sheet['B'+str(end+4)] = total_essential_repeat

# Saving the Excel file
workbook.save(save_file_at)

# Create empty dataframes to store rows
science_df = pd.DataFrame(columns= new_df.columns)
commerce_df = pd.DataFrame(columns= new_df.columns)
pcm_df = pd.DataFrame(columns=new_df.columns)
pcb_df = pd.DataFrame(columns=new_df.columns)

# Iterate through rows
for index, row in new_df.iterrows():
    if row['042'] != '':
        science_df = pd.concat([science_df, pd.DataFrame([row])], ignore_index=True)

        if row['041'] != '':       # DataFrame for PCM stream
            pcm_df = pd.concat([pcm_df, pd.DataFrame([row])], ignore_index=True)
        elif row['044'] != '':       # DataFrame for PCB stream
            pcb_df = pd.concat([pcb_df, pd.DataFrame([row])], ignore_index=True)

    elif row['030'] != '' or row['054'] != '' or row['055'] != '':
        commerce_df = pd.concat([commerce_df, pd.DataFrame([row])], ignore_index=True)

# Refining the columns one last time
for col in science_df.columns:
    if all(science_df[col] == ""):
        science_df.drop(columns=[col], inplace=True)

for col in commerce_df.columns:
    if all(commerce_df[col] == ""):
        commerce_df.drop(columns=[col], inplace=True)

for col in pcm_df.columns:
    if all(pcm_df[col] == ""):
        pcm_df.drop(columns=[col], inplace=True)
for col in pcb_df.columns:
    if all(pcb_df[col] == ""):
        pcb_df.drop(columns=[col], inplace=True)

# Acessing all the data to find QPI
sum_science = science_df['Total Marks (Best 5)'].sum()
len_science = len(science_df)
sum_commerce = commerce_df['Total Marks (Best 5)'].sum()
len_commerce = len(commerce_df)
t_sub = 5

# Finding QPI of PCM, PCB and Commerce :
qpi_science = (sum_science / (len_science * t_sub)) 
qpi_commerce = (sum_commerce /  (len_commerce * t_sub))

print(qpi_science)
print(qpi_commerce)
print(science_df)
print(commerce_df)

# Open the Excel file to add new sheets and data
with pd.ExcelWriter(save_file_at, engine='openpyxl', mode='a') as writer:
    science_df.to_excel(writer, sheet_name='Science', index=False)
    pcm_df.to_excel(writer, sheet_name='Maths', index=False)
    pcb_df.to_excel(writer, sheet_name='Biology', index=False)
    commerce_df.to_excel(writer, sheet_name='Commerce', index=False)

# Adding QPI in the Sheets :
workbook = load_workbook(save_file_at)
sheet_science = workbook['Science']     # For Science
end = sheet_science.max_row
sheet_science['A'+str(end+2)] = 'QPI : '
sheet_science['B'+str(end+2)] = qpi_science
workbook.save(save_file_at)

workbook = load_workbook(save_file_at)
sheet_commerce = workbook['Commerce']           # For Commerce
end = sheet_commerce.max_row
sheet_commerce['A'+str(end+2)] = 'QPI : '
sheet_commerce['B'+str(end+2)] = qpi_commerce
workbook.save(save_file_at)

# Counting the number of students Percentage wise :
count_above_90 = len(new_df[new_df['Percentage (%)'] >= 90])
count_between_80_and_89 = len(new_df[(new_df['Percentage (%)'] >= 80) & (new_df['Percentage (%)'] <= 89)])
count_between_75_and_79 = len(new_df[(new_df['Percentage (%)'] >= 75) & (new_df['Percentage (%)'] <= 79)])
count_between_60_and_74 = len(new_df[(new_df['Percentage (%)'] >= 60) & (new_df['Percentage (%)'] <= 74)])
count_below_60 = len(new_df[new_df['Percentage (%)'] < 60])

# Creating a dataframe for Percentage count : 
count_data = {'Percentage' : ['Above 90%', '80% to 89%', '75% to 79%', '60% to 74%', 'Below 60%'],
    'Total No of student' : [count_above_90, count_between_80_and_89, count_between_75_and_79, 
                                       count_between_60_and_74, count_below_60]}

percentage_count_df = pd.DataFrame(count_data)

print(percentage_count_df)

# Counting the number of Students Subject wise : 
Subject_percentage_count_df = pd.DataFrame({'Sr. No.':[1,2,3,4,5,6,7,8,9,10,11,12],
    'Subject Name':['Maths', 'Physics', 'Chemistry', 'Biology', 'English', 'Physical Education', 'Economics',
                    'Business Studies', 'Accountancy', 'Hindi', 'Sanskrit', 'I.P.']})

sub_codes = ['041','042','043','044','301','048','030','054','055','302','322','065']

#Empty lists to make columns
above_90 = []
between_80_89 = []
between_70_79 = []
between_60_69 = []
below_60 = []

# Iteration through every subject columns : 
for s_code in sub_codes:
    # Filter out empty strings before conversion
    non_empty_values = new_df[s_code][new_df[s_code] != '']
    non_empty_values = non_empty_values.astype(int, errors='ignore')
    
    count_above_90 = len(non_empty_values[non_empty_values >= 90])
    above_90.append(count_above_90)
    
    count_between_80_and_89 = len(non_empty_values[(non_empty_values >= 80) & (non_empty_values <= 89)])
    between_80_89.append(count_between_80_and_89)
    
    count_between_70_and_79 = len(non_empty_values[(non_empty_values >= 70) & (non_empty_values <= 79)])
    between_70_79.append(count_between_70_and_79)
    
    count_between_60_and_69 = len(non_empty_values[(non_empty_values >= 60) & (non_empty_values <= 69)])
    between_60_69.append(count_between_60_and_69)
    
    count_below_60 = len(non_empty_values[non_empty_values < 60])
    below_60.append(count_below_60)

Subject_percentage_count_df['90% and Above'] = above_90
Subject_percentage_count_df['80% To 89%'] = between_80_89
Subject_percentage_count_df['70% To 79%'] = between_70_79
Subject_percentage_count_df['60% To 69%'] = between_60_69
Subject_percentage_count_df['Below 60%'] = below_60

# Adding Total Students in a new column :
Subject_percentage_count_df['Total Student'] = Subject_percentage_count_df.iloc[:, 2:].sum(axis=1)

print(Subject_percentage_count_df)

highest_marks_df = pd.DataFrame({'Sr. No.':[1,2,3,4,5,6,7,8,9,10,11,12],
    'Subject Codes':['041','042','043','044','301','048','030','054','055','302','322','065'],
    'Subject Name':['Maths', 'Physics', 'Chemistry', 'Biology', 'English', 'Physical Education', 'Economics',
                    'Business Studies', 'Accountancy', 'Hindi', 'Sanskrit', 'I.P.']})

sub_codes = ['041','042','043','044','301','048','030','054','055','302','322','065']

# Create an empty list to store dictionaries of highest marks information
highest_marks_data = []

# Iterate through each subject
for s_code in sub_codes:
    # Convert the column to numeric (ignoring errors)
    new_df[s_code] = pd.to_numeric(new_df[s_code], errors='coerce')

    # Find the highest marks and the corresponding students' names
    highest_marks = new_df[s_code].max()
    highest_mark_students = new_df[new_df[s_code] == highest_marks]['Name'].tolist()
    highest_mark_students = ', '.join(highest_mark_students)  # Join names into a string and remove from list

    # Update the corresponding rows in the existing dataframe
    highest_marks_df.loc[highest_marks_df['Subject Codes'] == s_code, 'Highest Marks'] = highest_marks
    highest_marks_df.loc[highest_marks_df['Subject Codes'] == s_code, 'Name of Toppers'] = highest_mark_students

print(highest_marks_df)

# Load the existing Excel file
workbook = load_workbook(filename=save_file_at)
# Create a new sheet
sheet = workbook.create_sheet(title='Analysis')
sheet = workbook['Analysis']
sheet2 = workbook.create_sheet(title='Analysis 2')
sheet2 = workbook['Analysis 2']

# Writing DataFrame rows to Excel, starting from the specified row
for row in dataframe_to_rows(percentage_count_df, index=False, header=True):
    sheet.append(row)
sheet.append([' '])
for row in dataframe_to_rows(Subject_percentage_count_df, index=False, header=True, ):
    sheet.append(row)
sheet.append([' '])
for row in dataframe_to_rows(highest_marks_df, index=False, header=True, ):
    sheet.append(row)
sheet.append([' '])
for row in dataframe_to_rows(highest_marks_df, index=False, header=True, ):
    sheet2.append(row)
sheet2.append([' '])

# Saving the Excel file
workbook.save(save_file_at)

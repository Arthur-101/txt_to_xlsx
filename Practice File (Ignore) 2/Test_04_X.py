import re
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class DataProcessor:
    def __init__(self, input_file):
        self.input_file = input_file
        self.read_data()
        self.process_data()

    def read_data(self):
        with open(self.input_file) as file:
            self.reader = file.read()

    def process_data(self):
        date_pattern = r"DATE:- (\d{2}/\d{2}/\d{4})"
        school_pattern = r"SCHOOL : - (\d+) (.+)"
        region_pattern = r"REGION:\s+(\S+)"

        date_match = re.search(date_pattern, self.reader)
        school_match = re.search(school_pattern, self.reader)
        region_match = re.search(region_pattern, self.reader)

        self.date = date_match.group(1) if date_match else ""
        self.school_code = school_match.group(1) if school_match else ""
        self.school_name = school_match.group(2) if school_match else ""
        self.region = region_match.group(1) if region_match else ""

        total_candidates_pattern = r"TOTAL CANDIDATES\s*:\s*(\d+)"
        total_pass_pattern = r"TOTAL PASS\s*:\s*(\d+)"
        total_comptt_pattern = r"TOTAL COMPTT\.\s*:\s*(\d+)"
        total_essential_repeat_pattern = r"TOTAL ESSENTIAL REPEAT\s*:\s*(\d+)"
        total_absent_pattern = r"TOTAL ABSENT\s*:\s*(\d+)"

        total_candidates_match = re.search(total_candidates_pattern, self.reader)
        total_pass_match = re.search(total_pass_pattern, self.reader)
        total_comptt_match = re.search(total_comptt_pattern, self.reader)
        total_essential_repeat_match = re.search(total_essential_repeat_pattern, self.reader)
        total_absent_match = re.search(total_absent_pattern, self.reader)

        self.total_candidates = total_candidates_match.group(1) if total_candidates_match else ""
        self.total_pass = total_pass_match.group(1) if total_pass_match else ""
        self.total_comptt = total_comptt_match.group(1) if total_comptt_match else ""
        self.total_essential_repeat = total_essential_repeat_match.group(1) if total_essential_repeat_match else ""
        self.total_absent = total_absent_match.group(1) if total_absent_match else ""

        unwanted_pattern = r'DATE:.*?\n.*?-----.*?\n\nSCHOOL.*?\n'
        self.input_string_cleaned = re.sub(unwanted_pattern, '', self.reader, flags=re.DOTALL)

        roll_gender_name_pattern = r"(\d+)\s+(\w)\s+([A-Z ]+)"
        subject_codes_pattern = r"(\d{3}\s+){5}(\d{3})?"
        result_pattern = r"\b(PASS|FAIL|COMP|ABST)\b"
        marks_grades_pattern = r"(\d{3}|AB)\s+([A-Z]\d?)"

        self.students_data = []
        self.current_student_info = None
        self.current_student_grades = None
        lines = self.reader.strip().split('\n')

        for line in lines:
            line = line.strip()
            if re.match(roll_gender_name_pattern, line):
                self.current_student_info = line
            elif re.match(marks_grades_pattern, line):
                self.current_student_grades = line
                roll, gender, name = re.search(roll_gender_name_pattern, self.current_student_info).groups()
                subject_codes_string = re.search(subject_codes_pattern, self.current_student_info).group()
                subject_codes = re.findall(r"\d{3}", subject_codes_string)
                result = re.search(result_pattern, self.current_student_info).group()
                marks_grades = re.findall(marks_grades_pattern, self.current_student_grades)
                self.add_student_data(roll, gender, name, subject_codes, result, marks_grades)

        self.df = pd.DataFrame(self.students_data)
        self.df.reset_index(drop=True, inplace=True)
        self.df = self.df.fillna("")

        def extract_numeric(value):
            numeric_part = re.sub(r'\D', '', str(value))
            return int(numeric_part) if numeric_part.isdigit() else 0

        self.df['Marks_1'] = self.df['Marks_1'].apply(extract_numeric)
        self.df['Marks_2'] = self.df['Marks_2'].apply(extract_numeric)
        self.df['Marks_3'] = self.df['Marks_3'].apply(extract_numeric)
        self.df['Marks_4'] = self.df['Marks_4'].apply(extract_numeric)
        self.df['Marks_5'] = self.df['Marks_5'].apply(extract_numeric)
        self.df['Marks_6'] = self.df['Marks_6'].apply(extract_numeric)

        self.df['Total_Marks'] = self.df['Marks_1'] + self.df['Marks_2'] + self.df['Marks_3'] + self.df['Marks_4'] + self.df['Marks_5'] + self.df['Marks_6']

        self.df['Total Marks (Best 5)'] = self.df[['Marks_1', 'Marks_2', 'Marks_3', 'Marks_4', 'Marks_5', 'Marks_6']].apply(lambda row: sum(sorted(row, reverse=True)[:5]), axis=1)

        total_possible_marks = 500
        self.df['Percentage (%)'] = (self.df['Total Marks (Best 5)'] / total_possible_marks) * 100
        self.df['Percentage (%)'] = self.df['Percentage (%)'].apply(lambda x: round(x, 2))

        subject_codes = self.df[['Sub_1', 'Sub_2', 'Sub_3', 'Sub_4', 'Sub_5', 'Sub_6']].values.flatten()
        subject_codes = np.unique(subject_codes)

        new_columns = ['Roll', 'Gender', 'Name']
        for code in subject_codes:
            new_columns.append(code)
            new_columns.append('Grade_' + code)
        new_columns.append('Total Marks')
        new_columns.append('Total Marks (Best 5)')

        self.new_df = pd.DataFrame(columns=new_columns)

        for _, row in self.df.iterrows():
            new_row = [row['Roll'], row['Gender'], row['Name']]
            total_marks = 0
            best_marks = []
            for code in subject_codes:
                matching_index = np.where(row[['Sub_1', 'Sub_2', 'Sub_3', 'Sub_4', 'Sub_5', 'Sub_6']] == code)[0]
                if matching_index.size > 0:
                    marks = row['Marks_' + str(matching_index[0] + 1)]
                    grade = row['grade_' + str(matching_index[0] + 1)]
                    total_marks += marks
                    best_marks.append(marks)
                    new_row.append(marks)
                    new_row.append(grade)
                else:
                    new_row.append(np.NaN)
                    new_row.append(np.NaN)

            best_marks.sort(reverse=True)
            total_marks_best_5 = sum(best_marks[:5])
            new_row.append(total_marks)
            new_row.append(total_marks_best_5)

            self.new_df.loc[len(self.new_df)] = new_row

        total_possible_marks = 500
        self.new_df['Percentage (%)'] = (self.new_df['Total Marks (Best 5)'] / total_possible_marks) * 100
        self.new_df['Percentage (%)'] = self.new_df['Percentage (%)'].apply(lambda x: round(x, 2))

        for col in self.new_df.columns:
            if all(self.new_df[col] == ""):
                self.new_df.drop(columns=[col], inplace=True)

        columns_to_remove = [col for col in self.new_df.columns if col == ""]
        self.new_df.drop(columns=columns_to_remove, inplace=True)
        
        # Acessing all the data to find QPI
        sum_marks = self.new_df['Total Marks (Best 5)'].sum()
        len_students = len(self.new_df)
        t_sub = 5

        qpi_x = sum_marks / (len_students * t_sub)
        self.qpi_x = qpi_x

    def add_student_data(self, roll, gender, name, subject_codes, result, marks_grades):
        marks = []
        grades = []
        for mark_grade_tuple in marks_grades:
            mark, grade = mark_grade_tuple[0], mark_grade_tuple[1]
            if mark.isdigit():
                marks.append(int(mark))
            else:
                marks.append(mark)
            grades.append(grade)

        if len(subject_codes) < 6 and len(marks) < 6 and len(grades) < 6:
            subject_codes.append(np.NaN)
            marks.append(np.NaN)
            grades.append(np.NaN)

        row_data = {
            'Roll': roll,
            'Gender': gender,
            'Name': name.strip(),
            'Sub_1': subject_codes[0],
            'Marks_1': marks[0],
            'grade_1': grades[0],
            'Sub_2': subject_codes[1],
            'Marks_2': marks[1],
            'grade_2': grades[1],
            'Sub_3': subject_codes[2],
            'Marks_3': marks[2],
            'grade_3': grades[2],
            'Sub_4': subject_codes[3],
            'Marks_4': marks[3],
            'grade_4': grades[3],
            'Sub_5': subject_codes[4],
            'Marks_5': marks[4],
            'grade_5': grades[4],
            'Sub_6': subject_codes[5],
            'Marks_6': marks[5],
            'grade_6': grades[5],
            'Result': result
        }

        self.students_data.append(row_data)

    def calculate_percentage_counts(self):
        # Counting the number of students Percentage wise:
        count_above_90 = len(self.new_df[self.new_df['Percentage (%)'] >= 90])
        count_between_80_and_89 = len(self.new_df[(self.new_df['Percentage (%)'] >= 80) & (self.new_df['Percentage (%)'] <= 89)])
        count_between_75_and_79 = len(self.new_df[(self.new_df['Percentage (%)'] >= 75) & (self.new_df['Percentage (%)'] <= 79)])
        count_between_60_and_74 = len(self.new_df[(self.new_df['Percentage (%)'] >= 60) & (self.new_df['Percentage (%)'] <= 74)])
        count_below_60 = len(self.new_df[self.new_df['Percentage (%)'] < 60])

        count_data = {'Percentage': ['Above 90%', '80% to 89%', '75% to 79%', '60% to 74%', 'Below 60%'],
                      'Total No of student': [count_above_90, count_between_80_and_89, count_between_75_and_79,
                                               count_between_60_and_74, count_below_60]}
        self.percentage_count_df = pd.DataFrame(count_data)

        return self.percentage_count_df

    def calculate_subject_percentage_counts(self):
        # Counting the number of Students Subject wise:
        Subject_percentage_count_df = pd.DataFrame({'Sr. No.': [1, 2, 3, 4, 5, 6, 7],
                                                    'Subject Name': ['English', 'Hindi', 'Sanskrit', 'Maths', 'Science',
                                                                     'Social Science', 'I.T.']})

        sub_codes = ['184', '002', '122', '041', '086', '087', '402']

        above_90 = []
        between_80_89 = []
        between_70_79 = []
        between_60_69 = []
        below_60 = []

        for s_code in sub_codes:
            non_empty_values = self.new_df[s_code][self.new_df[s_code] != '']
            non_empty_values = non_empty_values.astype(int, errors='ignore')

            count_above_90 = len(non_empty_values[non_empty_values >= 90])
            above_90.append(count_above_90)

            count_between_80_and_89 = len(non_empty_values[(non_empty_values >= 80) & (non_empty_values <= 89)])
            between_80_89.append(count_between_80_and_89)

            count_between_70_and_79 = len(non_empty_values[(non_empty_values >= 70) & (non_empty_values <= 79)])
            between_70_79.append(count_between_70_and_79)

            count_between_60_and_69 = len(non_empty_values[(non_empty_values >= 60) & (non_empty_values <= 69)])
            between_60_69.append(count_between_60_and_69)

            count_below_60 = len(non_empty_values[non_empty_values < 60])
            below_60.append(count_below_60)

        Subject_percentage_count_df['90% and Above'] = above_90
        Subject_percentage_count_df['80% To 89%'] = between_80_89
        Subject_percentage_count_df['70% To 79%'] = between_70_79
        Subject_percentage_count_df['60% To 69%'] = between_60_69
        Subject_percentage_count_df['Below 60%'] = below_60

        Subject_percentage_count_df['Total Student'] = Subject_percentage_count_df.iloc[:, 2:].sum(axis=1)

        self.Subject_percentage_count_df = Subject_percentage_count_df
        
        return self.Subject_percentage_count_df

    def calculate_highest_marks_students(self):
        highest_marks_df = pd.DataFrame({'Sr. No.':[1,2,3,4,5,6,7],
            'Subject Codes':['184', '002', '122', '041', '086', '087', '402'],
            'Subject Name':['English', 'Hindi', 'Sanskrit', 'Maths', 'Science', 'Social Science', 'I.T.']})

        sub_codes = ['184', '002', '122', '041', '086', '087', '402']

        highest_marks_data = []

        for s_code in sub_codes:
            self.new_df[s_code] = pd.to_numeric(self.new_df[s_code], errors='coerce')
            highest_marks = self.new_df[s_code].max()
            highest_mark_students = self.new_df[self.new_df[s_code] == highest_marks]['Name'].tolist()
            highest_mark_students = ', '.join(highest_mark_students)
            highest_marks_df.loc[highest_marks_df['Subject Codes'] == s_code, 'Highest Marks'] = highest_marks
            highest_marks_df.loc[highest_marks_df['Subject Codes'] == s_code, 'Name of Toppers'] = highest_mark_students

        self.highest_marks_df = highest_marks_df
        
        return self.highest_marks_df

    def save_data_to_excel(self, output_file):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Result'

        sheet['A1'] = 'Date:-'
        sheet['A2'] = 'School Code'
        sheet['H1'] = 'Region:'
        sheet['B1'] = self.date
        sheet['B2'] = self.school_code
        sheet['I1'] = self.region

        start_column = 'C'
        end_column = 'G'
        merged_cell = sheet[start_column + '3']
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(f'{start_column}3:{end_column}3')
        sheet[f'{start_column}3'] = self.school_name

        for row in dataframe_to_rows(self.new_df, index=False, header=True):
            sheet.append(row)
        sheet.append([' '])

        start_row = 4
        end = sheet.max_row
        sheet['A'+str(end+2)] = 'Total Candidates :- '
        sheet['D'+str(end+2)] = 'Total Absent :- '
        sheet['A'+str(end+3)] = 'Total Pass :- '
        sheet['D'+str(end+3)] = 'Total Comptt. :-'
        sheet['A'+str(end+4)] = 'Total Essential Repeat :- '

        sheet['B'+str(end+2)] = self.total_candidates
        sheet['E'+str(end+2)] = self.total_absent
        sheet['B'+str(end+3)] = self.total_pass
        sheet['E'+str(end+3)] = self.total_comptt
        sheet['B'+str(end+4)] = self.total_essential_repeat
        
        # Adding QPI in the Sheets :
        sheet['A'+str(end+6)] = 'QPI : '
        sheet['B'+str(end+6)] = self.qpi_x

        workbook.save(output_file)
    
    def save_analysis_to_excel(self, output_file):
        # Load the existing Excel file
        workbook = load_workbook(filename=output_file)
        # Create a new sheet
        sheet = workbook.create_sheet(title='Analysis')
        sheet = workbook['Analysis']
        sheet2 = workbook.create_sheet(title='Analysis 2')
        sheet2 = workbook['Analysis 2']

        # Writing DataFrame rows to Excel, starting from the specified row
        for row in dataframe_to_rows(self.percentage_count_df, index=False, header=True):
            sheet.append(row)
        sheet.append([' '])

        for row in dataframe_to_rows(self.Subject_percentage_count_df, index=False, header=True, ):
            sheet.append(row)
        sheet.append([' '])

        for row in dataframe_to_rows(self.highest_marks_df, index=False, header=True, ):
            sheet.append(row)
        sheet.append([' '])

        for row in dataframe_to_rows(self.highest_marks_df, index=False, header=True, ):
            sheet2.append(row)
        sheet2.append([' '])

        # Saving the Excel file
        workbook.save(output_file)

if __name__ == "__main__":
    input_file = '65027 X.txt'
    output_file = 'Output_Excel_X.xlsx'
    data_processor = DataProcessor(input_file)
    data_processor.save_data_to_excel(output_file)
    data_processor.save_analysis_to_excel(output_file)

import re
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class DataProcessor_XII:
    # def __init__(self, input_file, output_file):
    #     self.input_file = input_file
    #     self.output_file = output_file
    #     self.read_data()
    #     self.process_data()
    #     self.calculate_qpi()
    #     self.calculate_percentage_counts()
    #     self.calculate_subject_percentage_counts()
    #     self.calculate_highest_marks()
    #     self.save_data_to_excel()
    def __init__(self, input_file):
        self.input_file = input_file
        self.read_data()
        self.process_data()

    def read_data(self):
        with open(self.input_file) as file:
            self.reader = file.read()
    
    def process_data(self):
        date_pattern = r"DATE:- (\d{2}/\d{2}/\d{4})"
        school_pattern = r"SCHOOL : - (\d+) (.+)"
        region_pattern = r"REGION:\s+(\S+)"

        date_match = re.search(date_pattern, self.reader)
        school_match = re.search(school_pattern, self.reader)
        region_match = re.search(region_pattern, self.reader)

        self.date = date_match.group(1) if date_match else ""
        self.school_code = school_match.group(1) if school_match else ""
        self.school_name = school_match.group(2) if school_match else ""
        self.region = region_match.group(1) if region_match else ""

        total_candidates_pattern = r"TOTAL CANDIDATES\s*:\s*(\d+)"
        total_pass_pattern = r"TOTAL PASS\s*:\s*(\d+)"
        total_comptt_pattern = r"TOTAL COMPTT\.\s*:\s*(\d+)"
        total_essential_repeat_pattern = r"TOTAL ESSENTIAL REPEAT\s*:\s*(\d+)"
        total_absent_pattern = r"TOTAL ABSENT\s*:\s*(\d+)"

        total_candidates_match = re.search(total_candidates_pattern, self.reader)
        total_pass_match = re.search(total_pass_pattern, self.reader)
        total_comptt_match = re.search(total_comptt_pattern, self.reader)
        total_essential_repeat_match = re.search(total_essential_repeat_pattern, self.reader)
        total_absent_match = re.search(total_absent_pattern, self.reader)

        self.total_candidates = total_candidates_match.group(1) if total_candidates_match else ""
        self.total_pass = total_pass_match.group(1) if total_pass_match else ""
        self.total_comptt = total_comptt_match.group(1) if total_comptt_match else ""
        self.total_essential_repeat = total_essential_repeat_match.group(1) if total_essential_repeat_match else ""
        self.total_absent = total_absent_match.group(1) if total_absent_match else ""

        unwanted_pattern = r'DATE:.*?\n.*?-----.*?\n\nSCHOOL.*?\n'
        self.input_string_cleaned = re.sub(unwanted_pattern, '', self.reader, flags=re.DOTALL)

        roll_gender_name_pattern = r"(\d+)\s+(\w)\s+([A-Z ]+)"
        subject_codes_pattern = r"(\d{3}\s+){5}(\d{3})?"
        result_pattern = r"\b(PASS|FAIL|COMP|ABST)\b"
        marks_grades_pattern = r"(\d{3}|AB)\s+([A-Z]\d?)"

        self.students_data = []
        self.current_student_info = None
        self.current_student_grades = None
        lines = self.reader.strip().split('\n')

        for line in lines:
            line = line.strip()
            if re.match(roll_gender_name_pattern, line):
                self.current_student_info = line
            elif re.match(marks_grades_pattern, line):
                self.current_student_grades = line
                roll, gender, name = re.search(roll_gender_name_pattern, self.current_student_info).groups()
                subject_codes_string = re.search(subject_codes_pattern, self.current_student_info).group()
                subject_codes = re.findall(r"\d{3}", subject_codes_string)
                result = re.search(result_pattern, self.current_student_info).group()
                marks_grades = re.findall(marks_grades_pattern, self.current_student_grades)
                self.add_student_data(roll, gender, name, subject_codes, result, marks_grades)

        self.df = pd.DataFrame(self.students_data)
        self.df.reset_index(drop=True, inplace=True)
        self.df = self.df.fillna("")

        def extract_numeric(value):
            numeric_part = re.sub(r'\D', '', str(value))
            return int(numeric_part) if numeric_part.isdigit() else 0

        self.df['Marks_1'] = self.df['Marks_1'].apply(extract_numeric)
        self.df['Marks_2'] = self.df['Marks_2'].apply(extract_numeric)
        self.df['Marks_3'] = self.df['Marks_3'].apply(extract_numeric)
        self.df['Marks_4'] = self.df['Marks_4'].apply(extract_numeric)
        self.df['Marks_5'] = self.df['Marks_5'].apply(extract_numeric)
        self.df['Marks_6'] = self.df['Marks_6'].apply(extract_numeric)

        self.df['Total_Marks'] = self.df['Marks_1'] + self.df['Marks_2'] + self.df['Marks_3'] + self.df['Marks_4'] + self.df['Marks_5'] + self.df['Marks_6']

        self.df['Total Marks (Best 5)'] = self.df[['Marks_1', 'Marks_2', 'Marks_3', 'Marks_4', 'Marks_5', 'Marks_6']].apply(lambda row: sum(sorted(row, reverse=True)[:5]), axis=1)

        total_possible_marks = 500
        self.df['Percentage (%)'] = (self.df['Total Marks (Best 5)'] / total_possible_marks) * 100
        self.df['Percentage (%)'] = self.df['Percentage (%)'].apply(lambda x: round(x, 2))

        subject_codes = self.df[['Sub_1', 'Sub_2', 'Sub_3', 'Sub_4', 'Sub_5', 'Sub_6']].values.flatten()
        subject_codes = np.unique(subject_codes)

        new_columns = ['Roll', 'Gender', 'Name']
        for code in subject_codes:
            new_columns.append(code)
            new_columns.append('Grade_' + code)
        new_columns.append('Total Marks')
        new_columns.append('Total Marks (Best 5)')

        self.new_df = pd.DataFrame(columns=new_columns)

        for _, row in self.df.iterrows():
            new_row = [row['Roll'], row['Gender'], row['Name']]
            total_marks = 0
            best_marks = []
            for code in subject_codes:
                matching_index = np.where(row[['Sub_1', 'Sub_2', 'Sub_3', 'Sub_4', 'Sub_5', 'Sub_6']] == code)[0]
                if matching_index.size > 0:
                    marks = row['Marks_' + str(matching_index[0] + 1)]
                    grade = row['grade_' + str(matching_index[0] + 1)]
                    total_marks += marks
                    best_marks.append(marks)
                    new_row.append(marks)
                    new_row.append(grade)
                else:
                    new_row.append(np.NaN)
                    new_row.append(np.NaN)

            best_marks.sort(reverse=True)
            total_marks_best_5 = sum(best_marks[:5])
            new_row.append(total_marks)
            new_row.append(total_marks_best_5)

            self.new_df.loc[len(self.new_df)] = new_row

        total_possible_marks = 500
        self.new_df['Percentage (%)'] = (self.new_df['Total Marks (Best 5)'] / total_possible_marks) * 100
        self.new_df['Percentage (%)'] = self.new_df['Percentage (%)'].apply(lambda x: round(x, 2))
        
        self.new_df = self.new_df.fillna('')

        for col in self.new_df.columns:
            if all(self.new_df[col] == ""):
                self.new_df.drop(columns=[col], inplace=True)

        columns_to_remove = [col for col in self.new_df.columns if col == ""]
        self.new_df.drop(columns=columns_to_remove, inplace=True)
        
        # Creating a df with renamed columns : 
        self.show_df = self.new_df.copy()   # copying the new_df first.
        self.show_df = self.show_df.fillna('')
                    # Create a mapping dictionary for column renaming
        column_mapping = {
            '041': 'Maths', 'Grade_041': 'Grade Maths', '042': 'Physics', 'Grade_042': 'Grade Physics',
            '043': 'Chemistry', 'Grade_043': 'Grade Chemistry', '044': 'Biology', 'Grade_044': 'Grade Biology',
            '301': 'English', 'Grade_301': 'Grade English', '048': 'Physicsl Education', 'Grade_048': 'Grade Physicsl Education',
            '030': 'Economics', 'Grade_030': 'Grade Economics', '054': 'Business Studies', 'Grade_054': 'Grade Business Studies',
            '055': 'Accountancy', 'Grade_055': 'Grade Accountancy', '302': 'Hindi', 'Grade_302': 'Grade Hindi',
            '322': 'Sanskrit', 'Grade_322': 'Grade Sanskrit', '065': 'I.P.', 'Grade_065': 'Grade I.P.'
            }
                    # Renaming the columns based on the mapping dictionary
        self.show_df.rename(columns=column_mapping, inplace=True)

        # Create a DataFrame for each stream by filtering the data
        self.science_df = self.show_df[self.show_df['Physics'] != '']            #  Science
        self.pcm_df = self.science_df[self.science_df['Maths'] != '']      #  PCM
        self.pcb_df = self.science_df[self.science_df['Biology'] != '']      #  PCB
        self.commerce_df = self.show_df[self.show_df['Economics'] != '']           #  Commerce

        # Reset the index of the new DataFrames
        self.science_df.reset_index(drop=True, inplace=True)
        self.pcm_df.reset_index(drop=True, inplace=True)
        self.pcb_df.reset_index(drop=True, inplace=True)
        self.commerce_df.reset_index(drop=True, inplace=True)

        # Refining the columns one last time
        for col in self.science_df.columns:
            if all(self.science_df[col] == ""):
                self.science_df.drop(columns=[col], inplace=True)

        for col in self.commerce_df.columns:
            if all(self.commerce_df[col] == ""):
                self.commerce_df.drop(columns=[col], inplace=True)

        for col in self.pcb_df.columns:
            if all(self.pcb_df[col] == ""):
                self.pcb_df.drop(columns=[col], inplace=True)
        
        # Creating a new DataFrame by removing all fields with Biology from Science df
        self.pcm_df = self.science_df[self.science_df['Biology'] == ""]
        self.pcm_df.reset_index(drop=True, inplace=True)

        # Removing empty columns
        for col in self.pcm_df.columns:
            if all(self.pcm_df[col] == ""):
                self.pcm_df.drop(columns=[col], inplace=True)


        # for col in self.pcm_df.columns:
        #     if all(self.pcm_df[col] == ""):
        #         self.pcm_df.drop(columns=[col], inplace=True)


        ####   QPI CALCULATIONS   ####
        # find QPI of Science
        sum_science = self.science_df['Total Marks (Best 5)'].sum()
        len_science = len(self.science_df)

        # find QPI of Commerce
        sum_commerce = self.commerce_df['Total Marks (Best 5)'].sum()
        len_commerce = len(self.commerce_df)

        # find QPI of PCM
        sum_pcm = self.pcm_df['Total Marks (Best 5)'].sum()
        len_pcm = len(self.pcm_df)

        # find QPI of PCB
        sum_pcb = self.pcb_df['Total Marks (Best 5)'].sum()
        len_pcb = len(self.pcb_df)

        # Finding QPI of PCM, PCB and Commerce :
        t_sub = 5
        qpi_science = (sum_science / (len_science * t_sub)) 
        qpi_commerce = (sum_commerce /  (len_commerce * t_sub))
        qpi_pcm = (sum_pcm / (len_pcm * t_sub))
        qpi_pcb = (sum_pcb / (len_pcb * t_sub))

        self.qpi_science = qpi_science
        self.qpi_commerce = qpi_commerce
        self.qpi_pcm = qpi_pcm
        self.qpi_pcb = qpi_pcb

    def add_student_data(self, roll, gender, name, subject_codes, result, marks_grades):
        marks = []
        grades = []
        for mark_grade_tuple in marks_grades:
            mark, grade = mark_grade_tuple[0], mark_grade_tuple[1]
            if mark.isdigit():
                marks.append(int(mark))
            else:
                marks.append(mark)
            grades.append(grade)

        if len(subject_codes) < 6 and len(marks) < 6 and len(grades) < 6:
            subject_codes.append(np.NaN)
            marks.append(np.NaN)
            grades.append(np.NaN)

        row_data = {
            'Roll': roll,
            'Gender': gender,
            'Name': name.strip(),
            'Sub_1': subject_codes[0],
            'Marks_1': marks[0],
            'grade_1': grades[0],
            'Sub_2': subject_codes[1],
            'Marks_2': marks[1],
            'grade_2': grades[1],
            'Sub_3': subject_codes[2],
            'Marks_3': marks[2],
            'grade_3': grades[2],
            'Sub_4': subject_codes[3],
            'Marks_4': marks[3],
            'grade_4': grades[3],
            'Sub_5': subject_codes[4],
            'Marks_5': marks[4],
            'grade_5': grades[4],
            'Sub_6': subject_codes[5],
            'Marks_6': marks[5],
            'grade_6': grades[5],
            'Result': result
        }

        self.students_data.append(row_data)

    def calculate_percentage_counts(self):
        # Counting the number of students Percentage wise:
        count_above_90 = len(self.new_df[self.new_df['Percentage (%)'] >= 90])
        count_between_80_and_89 = len(self.new_df[(self.new_df['Percentage (%)'] >= 80) & (self.new_df['Percentage (%)'] <= 89)])
        count_between_75_and_79 = len(self.new_df[(self.new_df['Percentage (%)'] >= 75) & (self.new_df['Percentage (%)'] <= 79)])
        count_between_60_and_74 = len(self.new_df[(self.new_df['Percentage (%)'] >= 60) & (self.new_df['Percentage (%)'] <= 74)])
        count_below_60 = len(self.new_df[self.new_df['Percentage (%)'] < 60])

        count_data = {'Percentage': ['Above 90%', '80% to 89%', '75% to 79%', '60% to 74%', 'Below 60%'],
                      'Total No of student': [count_above_90, count_between_80_and_89, count_between_75_and_79,
                                               count_between_60_and_74, count_below_60]}
        self.percentage_count_df = pd.DataFrame(count_data)

        return self.percentage_count_df

    def calculate_percentage_counts_plot(self):
        # Extract data for plotting
        percentage_labels = self.percentage_count_df['Percentage']
        student_counts = self.percentage_count_df['Total No of student']
         # Create a bar chart
        plt.figure(figsize=(8, 6))
        plt.bar(percentage_labels, student_counts, color='skyblue')
        plt.xlabel('Percentage Range')
        plt.ylabel('Total Number of Students')
        plt.title('Percentage-wise Student Counts')
        plt.yticks(np.arange(0, 140, 10))
        plt.xticks(rotation=15)  # Rotate x-axis labels for better readability
        
        plt.savefig('Items/xii_analysis_graph1.png')

    def calculate_subject_percentage_counts(self):
        # Counting the number of Students Subject wise:
        Subject_percentage_count_df = pd.DataFrame({'Sr. No.':[1,2,3,4,5,6,7,8,9,10,11,12],
                    'Subject Codes':['041','042','043','044','301','048','030','054','055','302','322','065'],
                    'Subject Name':['Maths', 'Physics', 'Chemistry', 'Biology','English', 'Physical Education',
                    'Economics', 'Business Studies', 'Accountancy', 'Hindi', 'Sanskrit', 'I.P.']})

        sub_codes = ['041','042','043','044','301','048','030','054','055','302','322','065']
        
        above_90 = []
        between_80_89 = []
        between_70_79 = []
        between_60_69 = []
        below_60 = []

        for s_code in sub_codes:
            non_empty_values = self.new_df[s_code][self.new_df[s_code] != '']
            non_empty_values = non_empty_values.astype(int, errors='ignore')

            count_above_90 = len(non_empty_values[non_empty_values >= 90])
            above_90.append(count_above_90)

            count_between_80_and_89 = len(non_empty_values[(non_empty_values >= 80) & (non_empty_values <= 89)])
            between_80_89.append(count_between_80_and_89)

            count_between_70_and_79 = len(non_empty_values[(non_empty_values >= 70) & (non_empty_values <= 79)])
            between_70_79.append(count_between_70_and_79)

            count_between_60_and_69 = len(non_empty_values[(non_empty_values >= 60) & (non_empty_values <= 69)])
            between_60_69.append(count_between_60_and_69)

            count_below_60 = len(non_empty_values[non_empty_values < 60])
            below_60.append(count_below_60)

        Subject_percentage_count_df['90% and Above'] = above_90
        Subject_percentage_count_df['80% To 89%'] = between_80_89
        Subject_percentage_count_df['70% To 79%'] = between_70_79
        Subject_percentage_count_df['60% To 69%'] = between_60_69
        Subject_percentage_count_df['Below 60%'] = below_60

        # Subject_percentage_count_df['Total Student'] = Subject_percentage_count_df.iloc[:, 2:].sum(axis=1)
        # Replace non-numeric values with NaN and then sum the columns
        Subject_percentage_count_df['Total Student'] = Subject_percentage_count_df.iloc[:, 2:].apply(pd.to_numeric, errors='coerce').sum(axis=1)

        self.Subject_percentage_count_df = Subject_percentage_count_df
        
        return self.Subject_percentage_count_df

    def calculate_subject_percentage_counts_plot(self):
        # Extract data for plotting
        subjects = self.Subject_percentage_count_df['Subject Name']
        percentage_categories = ['90% and Above', '80% To 89%', '70% To 79%', '60% To 69%', 'Below 60%']

        # Prepare data for plotting
        data = self.Subject_percentage_count_df[percentage_categories].transpose()

        # Create a bar chart for each subject
        fig, ax = plt.subplots(figsize=(10, 9))
        bar_width = 0.15
        index = range(len(subjects))

        for i, category in enumerate(percentage_categories):
            plt.bar([pos + i * bar_width for pos in index], data.loc[category], bar_width, label=category)

        plt.xlabel('Subjects')
        plt.ylabel('Total Number of Students')
        plt.title('Subject-wise Student Percentage Counts')
        plt.xticks([pos + bar_width * 2 for pos in index], subjects)
        plt.legend()
        plt.yticks(np.arange(0, 150, 10))
        # Rotate x-axis labels for better readability
        plt.xticks(rotation=25)
        
        plt.savefig('Items/xii_analysis_graph2.png')

    def calculate_highest_marks_students(self):
        highest_marks_df = pd.DataFrame({'Sr. No.':[1,2,3,4,5,6,7,8,9,10,11,12],
                    'Subject Codes':['041','042','043','044','301','048','030','054','055','302','322','065'],
                    'Subject Name':['Maths', 'Physics', 'Chemistry', 'Biology', 'English', 'Physical Education',
                    'Economics','Business Studies', 'Accountancy', 'Hindi', 'Sanskrit', 'I.P.']})

        sub_codes = ['041','042','043','044','301','048','030','054','055','302','322','065']

        highest_marks_data = []

        for s_code in sub_codes:
            self.new_df[s_code] = pd.to_numeric(self.new_df[s_code], errors='coerce')
            highest_marks = self.new_df[s_code].max()
            highest_mark_students = self.new_df[self.new_df[s_code] == highest_marks]['Name'].tolist()
            highest_mark_students = ', '.join(highest_mark_students)
            highest_marks_df.loc[highest_marks_df['Subject Codes'] == s_code, 'Highest Marks'] = highest_marks
            highest_marks_df.loc[highest_marks_df['Subject Codes'] == s_code, 'Name of Toppers'] = highest_mark_students

        self.highest_marks_df = highest_marks_df
        
        return self.highest_marks_df

    def calculate_highest_marks_students_plot(self):
        # Extract data for plotting
        subjects = self.highest_marks_df['Subject Name']
        highest_marks = self.highest_marks_df['Highest Marks']

        # Calculate the count of students who achieved the highest marks for each subject
        highest_marks_count = [len(names.split(', ')) for names in self.highest_marks_df['Name of Toppers']]

        # Create a bar chart for highest marks and the count of students with the highest marks
        fig, ax = plt.subplots(figsize=(10, 8))
        bar_width = 0.4
        index = range(len(subjects))

        plt.bar(index, highest_marks, bar_width, label='Highest Marks', color='skyblue')
        plt.bar([pos + bar_width for pos in index], highest_marks_count, bar_width, label='No. of Students', color='lightcoral')

        plt.xlabel('Subjects')
        plt.ylabel('Marks / No. of Students')
        plt.title('Highest Marks and No. of Students with Highest Marks for Each Subject')
        plt.xticks([pos + bar_width / 2 for pos in index], subjects)
        plt.legend()
        # Set y-axis tick marks at intervals of 20
        plt.yticks(np.arange(0, 105, 5))
        # Rotate x-axis labels for better readability
        plt.xticks(rotation=15)
        
        plt.savefig('Items/xii_analysis_graph3.png')

    def save_data_to_excel(self, output_file):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Result'

        # writing the headers
        sheet['A1'] = 'Date:-'
        sheet['A2'] = 'School Code'
        sheet['H1'] = 'Region:'
        sheet['B1'] = self.date
        sheet['B2'] = self.school_code
        sheet['I1'] = self.region

        # merging the headers
        start_column = 'C'
        end_column = 'G'
        merged_cell = sheet[start_column + '3']
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(f'{start_column}3:{end_column}3')
        sheet[f'{start_column}3'] = self.school_name

        # writing the data
        for row in dataframe_to_rows(self.show_df, index=False, header=True):
            sheet.append(row)
        sheet.append([' '])

        # adding the total counts
        start_row = 4
        end = sheet.max_row
        sheet['A'+str(end+2)] = 'Total Candidates :- '
        sheet['D'+str(end+2)] = 'Total Absent :- '
        sheet['A'+str(end+3)] = 'Total Pass :- '
        sheet['D'+str(end+3)] = 'Total Comptt. :-'
        sheet['A'+str(end+4)] = 'Total Essential Repeat :- '

        sheet['B'+str(end+2)] = self.total_candidates
        sheet['E'+str(end+2)] = self.total_absent
        sheet['B'+str(end+3)] = self.total_pass
        sheet['E'+str(end+3)] = self.total_comptt
        sheet['B'+str(end+4)] = self.total_essential_repeat


        workbook.save(output_file)
        
        # Open the Excel file to add new sheets and data
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
            self.science_df.to_excel(writer, sheet_name='Science', index=False)
            self.pcm_df.to_excel(writer, sheet_name='Maths', index=False)
            self.pcb_df.to_excel(writer, sheet_name='Biology', index=False)
            self.commerce_df.to_excel(writer, sheet_name='Commerce', index=False)

        # Adding QPI in the Sheets :
        workbook = load_workbook(output_file)
        sheet_science = workbook['Science']     # For Science
        end = sheet_science.max_row
        sheet_science['A'+str(end+2)] = 'QPI : '
        sheet_science['B'+str(end+2)] = self.qpi_science
        workbook.save(output_file)

        workbook = load_workbook(output_file)
        sheet_commerce = workbook['Commerce']           # For Commerce
        end = sheet_commerce.max_row
        sheet_commerce['A'+str(end+2)] = 'QPI : '
        sheet_commerce['B'+str(end+2)] = self.qpi_commerce
        workbook.save(output_file)

        workbook = load_workbook(output_file)
        sheet_pcb = workbook['Biology']                 # For Biology
        end = sheet_pcb.max_row
        sheet_pcb['A'+str(end+2)] = 'QPI : '
        sheet_pcb['B'+str(end+2)] = self.qpi_pcb
        workbook.save(output_file)

        workbook = load_workbook(output_file)
        sheet_pcm = workbook['Maths']                   # For Maths
        end = sheet_pcm.max_row
        sheet_pcm['A'+str(end+2)] = 'QPI : '
        sheet_pcm['B'+str(end+2)] = self.qpi_pcm
        workbook.save(output_file)

    def save_analysis_to_excel(self, output_file):
        # Load the existing Excel file
        workbook = load_workbook(filename=output_file)
        # Create a new sheet
        sheet = workbook.create_sheet(title='Analysis 1')
        sheet = workbook['Analysis 1']
        sheet2 = workbook.create_sheet(title='Analysis 2')
        sheet2 = workbook['Analysis 2']

        # Writing DataFrame rows to Excel, starting from the specified row
        for row in dataframe_to_rows(self.percentage_count_df, index=False, header=True):
            sheet.append(row)
        sheet.append([' '])

        for row in dataframe_to_rows(self.Subject_percentage_count_df, index=False, header=True, ):
            sheet.append(row)
        sheet.append([' '])

        for row in dataframe_to_rows(self.highest_marks_df, index=False, header=True, ):
            sheet.append(row)
        sheet.append([' '])

        for row in dataframe_to_rows(self.highest_marks_df, index=False, header=True, ):
            sheet2.append(row)
        sheet2.append([' '])

        # Saving the Excel file
        workbook.save(output_file)


# if __name__ == "__main__":
#     input_file = '65027 XII.txt'
#     output_file = 'Output_Excel_2.xlsx'
#     data_processor = DataProcessor_XII(input_file, output_file)
if __name__ == "__main__":
    input_file = '65027 X.txt'
    output_file = 'Output_Excel_X.xlsx'
    data_processor = DataProcessor_XII(input_file)
    data_processor.save_data_to_excel(output_file)
    data_processor.save_analysis_to_excel(output_file)
